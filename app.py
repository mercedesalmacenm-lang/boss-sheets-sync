import os
import re
import json
import time
import unicodedata
import threading
import gspread
from google.oauth2.service_account import Credentials
import openpyxl
from datetime import datetime
from flask import Flask, request, jsonify, render_template_string

app = Flask(__name__)

SHEET_ID = os.environ.get("SHEET_ID", "1FLznJQ0PBxqnMNRPI_JEgv_QD7o7RoiodZLZmDzGE6Y")
desktop = os.path.join(os.path.expanduser("~"), "Desktop")
UPLOAD_DIR = os.environ.get("UPLOAD_DIR", desktop if os.path.isdir(desktop) else "./uploads")
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

os.makedirs(UPLOAD_DIR, exist_ok=True)

def get_credentials():
    creds_json = os.environ.get("GOOGLE_CREDENTIALS")
    if creds_json:
        creds_info = json.loads(creds_json)
        return Credentials.from_service_account_info(creds_info, scopes=SCOPES)
    creds_file = os.environ.get("GOOGLE_CREDENTIALS_FILE", "./credentials.json")
    return Credentials.from_service_account_file(creds_file, scopes=SCOPES)

def get_sheet():
    creds = get_credentials()
    client = gspread.authorize(creds)
    return client.open_by_key(SHEET_ID)

ENCABEZADOS_REQUERIDOS = ["articulo", "descripcion", "ubicacion"]

def normalizar(texto):
    import unicodedata
    texto = str(texto).lower().strip()
    texto = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in texto if not unicodedata.combining(c))

def es_archivo_boss(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(max_row=10, values_only=True):
            rows.append([str(c).strip() if c is not None else "" for c in row])

        tiene_titulo = any("maquinas fer" in normalizar(r[0]) for r in rows[:3] if r[0])
        tiene_inventarios = any("inventarios" in normalizar(r[0]) for r in rows[:3] if r[0])
        tiene_fecha = any(__import__("re").search(r'\d{1,2}/\w+/\d{2}', r[0]) for r in rows[:4] if r[0])

        tiene_encabezados = False
        for r in rows:
            vals = [normalizar(v) for v in r if v]
            if any(h in v for v in vals for h in ENCABEZADOS_REQUERIDOS):
                tiene_encabezados = True
                break

        wb.close()
        return tiene_titulo and tiene_inventarios and tiene_fecha and tiene_encabezados
    except Exception:
        pass
    return False

def read_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            data.append([str(cell) if cell is not None else "" for cell in row])
    wb.close()
    return data

def upload_to_sheet(data, sheet, filename):
    try:
        worksheets = sheet.worksheets()
        ws = worksheets[0] if worksheets else sheet.add_worksheet(
            title='BOSS', rows=len(data)+10, cols=len(data[0]) if data else 10
        )

        if not data:
            return True, 0

        headers = [str(c).lower().strip() for c in data[0]]
        almacen_idx = None
        for i, h in enumerate(headers):
            if normalizar(h) == "almacen":
                almacen_idx = i
                break

        if almacen_idx is not None and len(data) > 1:
            all_data = ws.get_all_values()
            if all_data:
                existing_headers = [str(c).lower().strip() for c in all_data[0]]
                existing_almacen_idx = None
                for i, h in enumerate(existing_headers):
                    if normalizar(h) == "almacen":
                        existing_almacen_idx = i
                        break

                if existing_almacen_idx is not None:
                    nuevo_almacen = str(data[1][almacen_idx]).strip().upper()
                    kept = [all_data[0]]
                    for row in all_data[1:]:
                        if len(row) > existing_almacen_idx:
                            if str(row[existing_almacen_idx]).strip().upper() != nuevo_almacen:
                                kept.append(row)
                    new_rows = [r for r in data[1:] if any(cell for cell in r)]
                    final = kept + new_rows
                    ws.clear()
                    ws.update(range_name='A1', values=final)
                    return True, len(new_rows)

        ws.clear()
        ws.update(range_name='A1', values=data)
        return True, len(data) - 1
    except Exception as e:
        return False, str(e)

def process_file(filepath, filename, sheet):
    if es_archivo_boss(filepath):
        data = read_excel(filepath)
        if data:
            success, result = upload_to_sheet(data, sheet, filename)
            return {
                "file": filename,
                "uploaded": success,
                "rows": result if success else None,
                "error": result if not success else None,
                "timestamp": datetime.now().isoformat()
            }
    return None

upload_log = []

def monitor_loop():
    processed = []
    while True:
        try:
            sheet = get_sheet()
            files = [f for f in os.listdir(UPLOAD_DIR) if f.endswith('.xlsx') and not f.startswith('~$')]

            for filename in files:
                if filename not in processed:
                    filepath = os.path.join(UPLOAD_DIR, filename)
                    result = process_file(filepath, filename, sheet)
                    if result:
                        upload_log.append(result)
                        processed.append(filename)
                        print(f"[UPLOAD] {filename}: {result}")

            time.sleep(10)
        except Exception as e:
            print(f"[ERROR] {e}")
            time.sleep(30)

@app.route("/")
def index():
    return render_template_string(PAGINA_SUBIDA)

@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "No se envió archivo"}), 400
    f = request.files["file"]
    if not f.filename.endswith(".xlsx"):
        return jsonify({"error": "Solo archivos .xlsx"}), 400
    filepath = os.path.join(UPLOAD_DIR, f.filename)
    f.save(filepath)
    sheet = get_sheet()
    result = process_file(filepath, f.filename, sheet)
    if result:
        upload_log.append(result)
        return jsonify(result)
    return jsonify({"error": "El archivo no tiene formato BOSS"}), 400

@app.route("/api/upload", methods=["POST"])
def api_upload():
    return upload()

@app.route("/status")
def status():
    return jsonify({
        "status": "active",
        "uploads": upload_log[-20:]
    })

@app.route("/health")
def health():
    return jsonify({"status": "ok"}), 200

PAGINA_SUBIDA = """
<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>BOSS Sync - Subir archivo</title>
<style>
  *{box-sizing:border-box;margin:0;padding:0}
  body{font-family:-apple-system,Segoe UI,Roboto,sans-serif;background:#181a1d;color:#ececea;min-height:100vh;display:flex;align-items:center;justify-content:center}
  .card{background:#212327;border:1px solid #3a3d43;border-radius:12px;padding:32px;max-width:480px;width:100%}
  h1{color:#f5a623;font-size:22px;margin-bottom:4px}
  .sub{color:#8b8d92;font-size:14px;margin-bottom:24px}
  .upload-area{border:2px dashed #3a3d43;border-radius:10px;padding:40px 20px;text-align:center;cursor:pointer;transition:border-color .15s}
  .upload-area:hover,.upload-area.dragover{border-color:#f5a623}
  .upload-area p{color:#8b8d92;font-size:14px}
  .upload-area .icon{font-size:36px;margin-bottom:8px}
  #fileInput{display:none}
  .btn{background:#f5a623;color:#181a1d;border:none;border-radius:8px;padding:12px 24px;font-size:15px;font-weight:700;cursor:pointer;width:100%;margin-top:16px;display:none}
  .btn:hover{opacity:.9}
  .result{margin-top:16px;padding:12px;border-radius:8px;font-size:14px;display:none}
  .result.ok{background:#1a2e1a;color:#5aa96a;border:1px solid #2a4a2a}
  .result.err{background:#2e1a1a;color:#c1443c;border:1px solid #4a2a2a}
  .log{margin-top:24px}
  .log h3{font-size:14px;color:#8b8d92;margin-bottom:8px;text-transform:uppercase;letter-spacing:.06em}
  .log-item{padding:8px 12px;background:#181a1d;border-radius:6px;margin-bottom:6px;font-size:13px;display:flex;justify-content:space-between}
  .log-item .ok{color:#5aa96a}
  .log-item .err{color:#c1443c}
</style>
</head>
<body>
<div class="card">
  <h1>BOSS Sheets Sync</h1>
  <p class="sub">Sube tu export de inventario desde BOSS</p>
  <div class="upload-area" id="dropZone" onclick="document.getElementById('fileInput').click()">
    <div class="icon">&#128228;</div>
    <p id="dropText">Arrastra tu archivo .xlsx aquí<br>o haz clic para seleccionar</p>
  </div>
  <input type="file" id="fileInput" accept=".xlsx">
  <button class="btn" id="uploadBtn">Subir archivo</button>
  <div class="result" id="result"></div>
  <div class="log" id="logSection">
    <h3>Últimas subidas</h3>
    <div id="logList"></div>
  </div>
</div>
<script>
const dropZone=document.getElementById('dropZone');
const fileInput=document.getElementById('fileInput');
const uploadBtn=document.getElementById('uploadBtn');
const result=document.getElementById('result');
let selectedFile=null;

fileInput.addEventListener('change',e=>{if(e.target.files[0])selectFile(e.target.files[0])});
dropZone.addEventListener('dragover',e=>{e.preventDefault();dropZone.classList.add('dragover')});
dropZone.addEventListener('dragleave',()=>dropZone.classList.remove('dragover'));
dropZone.addEventListener('drop',e=>{e.preventDefault();dropZone.classList.remove('dragover');if(e.dataTransfer.files[0])selectFile(e.dataTransfer.files[0])});

function selectFile(file){
  selectedFile=file;
  document.getElementById('dropText').innerHTML='<strong>'+file.name+'</strong><br>'+Math.round(file.size/1024)+' KB';
  uploadBtn.style.display='block';
  result.style.display='none';
}

uploadBtn.addEventListener('click',async()=>{
  if(!selectedFile)return;
  uploadBtn.textContent='Subiendo...';
  uploadBtn.disabled=true;
  const fd=new FormData();
  fd.append('file',selectedFile);
  try{
    const r=await fetch('/api/upload',{method:'POST',body:fd});
    const d=await r.json();
    result.style.display='block';
    if(d.error){result.className='result err';result.textContent=d.error}
    else{result.className='result ok';result.textContent='Subido: '+d.rows+' filas de datos'}
  }catch(e){result.style.display='block';result.className='result err';result.textContent='Error de conexión'}
  uploadBtn.textContent='Subir archivo';
  uploadBtn.disabled=false;
  loadLog();
});

async function loadLog(){
  try{
    const r=await fetch('/status');
    const d=await r.json();
    const list=document.getElementById('logList');
    if(!d.uploads||!d.uploads.length){list.innerHTML='<p style="color:#8b8d92;font-size:13px">Sin subidas aún</p>';return}
    list.innerHTML=d.uploads.reverse().map(u=>'<div class="log-item"><span>'+u.file+'</span><span class="'+(u.uploaded?'ok':'err')+'">'+(u.uploaded?u.rows+' filas':u.error)+'</span></div>').join('');
  }catch(e){}
}
loadLog();
</script>
</body>
</html>
"""

if __name__ == "__main__":
    monitor_thread = threading.Thread(target=monitor_loop, daemon=True)
    monitor_thread.start()
    print("[STARTED] Monitor BOSS Sheets Sync")
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
