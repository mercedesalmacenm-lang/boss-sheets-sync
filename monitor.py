import os
import re
import json
import time
import unicodedata
import gspread
from google.oauth2.service_account import Credentials
import openpyxl
from datetime import datetime

SHEET_ID = os.environ.get("SHEET_ID", "1FLznJQ0PBxqnMNRPI_JEgv_QD7o7RoiodZLZmDzGE6Y")
desktop = os.path.join(os.path.expanduser("~"), "Desktop")
UPLOAD_DIR = os.environ.get("UPLOAD_DIR", desktop if os.path.isdir(desktop) else "./uploads")
SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]

os.makedirs(UPLOAD_DIR, exist_ok=True)

def get_credentials():
    creds_json = os.environ.get("GOOGLE_CREDENTIALS")
    if creds_json:
        creds_info = json.loads(creds_json)
        return Credentials.from_service_account_info(creds_info, scopes=SCOPES)
    creds_file = os.environ.get("GOOGLE_CREDENTIALS_FILE", "./credentials.json")
    return Credentials.from_service_account_file(creds_file, scopes=SCOPES)

def get_sheet():
    creds = get_credentials()
    client = gspread.authorize(creds)
    return client.open_by_key(SHEET_ID)

def normalizar(texto):
    texto = str(texto).lower().strip()
    texto = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in texto if not unicodedata.combining(c))

def es_archivo_boss(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(max_row=10, values_only=True):
            rows.append([str(c).strip() if c is not None else "" for c in row])

        tiene_titulo = any("maquinas fer" in normalizar(r[0]) for r in rows[:3] if r[0])
        tiene_inventarios = any("inventarios" in normalizar(r[0]) for r in rows[:3] if r[0])
        tiene_fecha = any(re.search(r'\d{1,2}/\w+/\d{2}', r[0]) for r in rows[:4] if r[0])

        tiene_encabezados = False
        for r in rows:
            vals = [normalizar(v) for v in r if v]
            if any(h in v for v in vals for h in ["articulo", "descripcion", "ubicacion"]):
                tiene_encabezados = True
                break

        wb.close()
        return tiene_titulo and tiene_inventarios and tiene_fecha and tiene_encabezados
    except Exception:
        pass
    return False

def read_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            data.append([str(cell) if cell is not None else "" for cell in row])
    wb.close()
    return data

def detectar_almacen(data):
    if not data or len(data) < 2:
        return None
    headers = [normalizar(str(c)) for c in data[0]]
    for i, h in enumerate(headers):
        if h == "almacen":
            for row in data[1:]:
                if i < len(row) and str(row[i]).strip():
                    return str(row[i]).strip().upper()
    return None

def subir_a_pestana(data, sheet, nombre_pestana):
    try:
        worksheets = sheet.worksheets()
        ws = None
        for ws in worksheets:
            if ws.title.lower() == nombre_pestana.lower():
                break
        else:
            ws = sheet.add_worksheet(title=nombre_pestana, rows=len(data)+10, cols=len(data[0]) if data else 10)

        ws.clear()
        ws.update(range_name='A1', values=data)
        return True, len(data) - 1
    except Exception as e:
        return False, str(e)

def process_file(filepath, filename, sheet):
    if not es_archivo_boss(filepath):
        return None

    data = read_excel(filepath)
    if not data:
        return None

    almacen = detectar_almacen(data)
    if not almacen:
        print(f"[WARN] {filename}: No se detectó almacén, usando 'GENERAL'")
        almacen = "GENERAL"

    success, result = subir_a_pestana(data, sheet, almacen)
    return {
        "file": filename,
        "almacen": almacen,
        "uploaded": success,
        "rows": result if success else None,
        "error": result if not success else None,
        "timestamp": datetime.now().isoformat()
    }

def main():
    print(f"[STARTED] Monitor BOSS Sheets Sync")
    print(f"[INFO] Escaneando: {UPLOAD_DIR}")
    print(f"[INFO] Google Sheet: {SHEET_ID}")
    print()

    processed = []
    while True:
        try:
            sheet = get_sheet()
            files = [f for f in os.listdir(UPLOAD_DIR) if f.endswith('.xlsx') and not f.startswith('~$')]

            for filename in files:
                if filename not in processed:
                    filepath = os.path.join(UPLOAD_DIR, filename)
                    result = process_file(filepath, filename, sheet)
                    if result:
                        status = "OK" if result["uploaded"] else "ERROR"
                        print(f"[{status}] {result['almacen']}: {filename} -> {result.get('rows', result.get('error'))} filas")
                        processed.append(filename)

            time.sleep(10)
        except Exception as e:
            print(f"[ERROR] {e}")
            time.sleep(30)

if __name__ == "__main__":
    main()
